# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm
from tkinter import Tk, ttk
from tkinter.messagebox import showinfo

def ECD_PIS(ecd_pis_path, dado, batch_size=20):
    data = ''

    def FormatarData(variavel, num):
        global data
        data = variavel[num][:2] + '/' + variavel[num][2:4] + '/' + variavel[num][4:]
        return data

    ecd_pis_list = ecd_pis_path

    # Criar um workbook vazio
    wb = Workbook()

    # Criar duas planilhas vazias
    ws_a = wb.create_sheet("0000")
    ws_b = wb.create_sheet("0150")

    # Criar dataframes vazios
    df_0000 = pd.DataFrame()
    df_0001 = pd.DataFrame()

    # Função para atualizar a barra de progresso
    def update_progress(count):
        progress_bar["value"] = count
        root.update_idletasks()

    # Função para mostrar mensagem ao concluir
    def show_completion_message():
        showinfo("Concluído", "Processamento concluído com sucesso!")

    # Configurar janela
    root = Tk()
    root.title("Processando arquivos")
    root.geometry("300x100")

    # Criar barra de progresso
    progress_bar = ttk.Progressbar(root, length=200, mode="determinate")
    progress_bar.pack(pady=20)

    # Percorrer a lista de diretórios/arquivos em lotes com barra de progresso
    with tqdm(total=len(ecd_pis_list), desc="Processando arquivos", bar_format="{desc}") as pbar:
        for i in range(0, len(ecd_pis_list), batch_size):
            batch_paths = ecd_pis_list[i:i + batch_size]

            for path in batch_paths:
                try:
                    replace_ecd_pis_str = path.replace("/", "\\")

                    with open(replace_ecd_pis_str, 'r', encoding="latin-1") as file:
                        # Bloco 0
                        cont_0000 = {}
                        cont_0001 = {}

                        for line in file:
                            first_six_chars = line[:6]

                            if '|0000|' in first_six_chars:
                                # Split the line by '|' and store the resulting columns in a list
                                columns_0000 = line.strip().split('|')
                                # assign to every column
                                # cont_0000['ID_DT_INI'] = data
                                # FormatarData(columns_0000, 7)
                                # cont_0000['ID_CNPJ'] = columns_0000.get(9, '') para teste caso for nulo futuramente
                                cont_0000['REG'] = columns_0000[1]
                                cont_0000['COD_VER'] = columns_0000[2]
                                cont_0000['TIPO_ESCRIT'] = columns_0000[3]
                                cont_0000['IND_SIT_ESP'] = columns_0000[4]
                                cont_0000['NUM_REC_ANTERIOR'] = columns_0000[5]
                                cont_0000['DT_INI'] = columns_0000[6]
                                cont_0000['DT_FIN'] = columns_0000[7]
                                cont_0000['NOME'] = columns_0000[8]
                                cont_0000['CNPJ'] = columns_0000[9]
                                cont_0000['UF'] = columns_0000[10]
                                cont_0000['COD_MUN'] = columns_0000[11]
                                cont_0000['SUFRAMA'] = columns_0000[12]
                                cont_0000['IND_NAT_PJ'] = columns_0000[13]
                                cont_0000['IND_ATIV'] = columns_0000[14]
                                # Create a new row in the dataframe using the columns list as the values
                                row_0000 = pd.DataFrame(cont_0000, index=[0])
                                # Concatenate the row to the dataframe
                                df_0000 = pd.concat([df_0000, row_0000])
                            elif '|0001|' in first_six_chars:
                                # Split the line by '|' and store the resulting columns in a list
                                columns_0001 = line.strip().split('|')
                                cont_0001['REG'] = columns_0001[1]
                                cont_0001['IND_MOV'] = columns_0001[2]
                                row_0001 = pd.DataFrame(cont_0001, index=[0])
                                df_0001 = pd.concat([df_0001, row_0001])


                except FileNotFoundError as e:
                    print(f"Arquivo não encontrado: {path}")
                    return
                except PermissionError as e:
                    print(f"Erro de permissão ao acessar o arquivo: {path}")
                    return
                except IOError as e:
                    print(f"Erro de E/S ao abrir o arquivo: {path}")
                    return

                pbar.update(1)  # Atualiza a barra de progresso
                update_progress(pbar.n)  # Atualiza a barra de progresso da janela

    # Mostrar mensagem de conclusão
    show_completion_message()

    # Fechar a janela
    root.destroy()

    if dado == '1':
        # if dado equal to 1 I need to send the df_0000 and df_0150 to parte_pandas
        from ..main import MyApp
        app = MyApp.get_running_app()
        get_all_df = [df_0000, df_0001]
        # 0 = efd icms, 1 = efd pis, 2 = ecd icms, 3 = ecd pis
        type = '0'
        app.receive_df(get_all_df, type)
        return True
    else:
        for r in dataframe_to_rows(df_0000, index=False, header=True):
            ws_a.append(r)

        for r in dataframe_to_rows(df_0001, index=False, header=True):
            ws_b.append(r)

        first_sheet = wb.sheetnames[0]
        if first_sheet != ws_a.title and first_sheet != ws_b.title:
            sheet = wb[first_sheet]
            wb.remove(sheet)

        wb.save("D:\\Documents\\app\\test_ecd_PIS.xlsx")
