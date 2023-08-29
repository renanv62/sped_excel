import sys
import pandas as pd
import glob
import datetime
import csv
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm
from tkinter import Tk, ttk
from tkinter.messagebox import showinfo


def EFD_ICMS(icms_path, dado, batch_size=20):
    data = ''

    def FormatarData(variavel, num):
        global data
        data = variavel[num][:2] + '/' + variavel[num][2:4] + '/' + variavel[num][4:]
        return data

    icms_list = icms_path

    # Criar um workbook vazio
    wb = Workbook()

    # Criar duas planilhas vazias
    ws_a = wb.create_sheet("0000")
    ws_b = wb.create_sheet("0150")

    # Criar dataframes vazios
    # Bloco 0
    df_0000 = pd.DataFrame()
    df_0001 = pd.DataFrame()
    df_0002 = pd.DataFrame()
    df_0005 = pd.DataFrame()
    df_0015 = pd.DataFrame()
    df_0100 = pd.DataFrame()
    df_0150 = pd.DataFrame()
    df_0175 = pd.DataFrame()
    df_0190 = pd.DataFrame()
    df_0200 = pd.DataFrame()
    df_0205 = pd.DataFrame()
    df_0206 = pd.DataFrame()
    df_0210 = pd.DataFrame()
    df_0220 = pd.DataFrame()
    df_0221 = pd.DataFrame()
    df_0300 = pd.DataFrame()
    df_0305 = pd.DataFrame()
    df_0400 = pd.DataFrame()
    df_0450 = pd.DataFrame()
    df_0460 = pd.DataFrame()
    df_0500 = pd.DataFrame()
    df_0600 = pd.DataFrame()
    df_0990 = pd.DataFrame()
    # Bloco B
    df_B001 = pd.DataFrame()
    df_B020 = pd.DataFrame()
    df_B025 = pd.DataFrame()
    df_B030 = pd.DataFrame()
    df_B035 = pd.DataFrame()
    df_B350 = pd.DataFrame()
    df_B420 = pd.DataFrame()
    df_B440 = pd.DataFrame()
    df_B460 = pd.DataFrame()
    df_B470 = pd.DataFrame()
    df_B500 = pd.DataFrame()
    df_B510 = pd.DataFrame()
    df_B990 = pd.DataFrame()
    # Bloco C
    df_C001 = pd.DataFrame()
    df_C100 = pd.DataFrame()
    df_C101 = pd.DataFrame()
    df_C105 = pd.DataFrame()
    df_C110 = pd.DataFrame()
    df_C111 = pd.DataFrame()
    df_C112 = pd.DataFrame()
    df_C113 = pd.DataFrame()
    df_C114 = pd.DataFrame()
    df_C115 = pd.DataFrame()
    df_C116 = pd.DataFrame()
    df_C120 = pd.DataFrame()
    df_C130 = pd.DataFrame()
    df_C140 = pd.DataFrame()
    df_C141 = pd.DataFrame()
    df_C160 = pd.DataFrame()
    df_C165 = pd.DataFrame()
    df_C170 = pd.DataFrame()
    df_C171 = pd.DataFrame()
    df_C172 = pd.DataFrame()
    df_C173 = pd.DataFrame()
    df_C174 = pd.DataFrame()
    df_C175 = pd.DataFrame()
    df_C176 = pd.DataFrame()
    df_C177 = pd.DataFrame()
    df_C178 = pd.DataFrame()
    df_C179 = pd.DataFrame()
    df_C180 = pd.DataFrame()
    df_C181 = pd.DataFrame()
    df_C185 = pd.DataFrame()
    df_C186 = pd.DataFrame()
    df_C190 = pd.DataFrame()
    df_C191 = pd.DataFrame()
    df_C195 = pd.DataFrame()
    df_C197 = pd.DataFrame()
    df_C300 = pd.DataFrame()
    df_C310 = pd.DataFrame()
    df_C320 = pd.DataFrame()
    df_C321 = pd.DataFrame()
    df_C330 = pd.DataFrame()
    df_C350 = pd.DataFrame()
    df_C370 = pd.DataFrame()
    df_C380 = pd.DataFrame()
    df_C390 = pd.DataFrame()
    df_C400 = pd.DataFrame()
    df_C405 = pd.DataFrame()
    df_C410 = pd.DataFrame()
    df_C420 = pd.DataFrame()
    df_C425 = pd.DataFrame()
    df_C430 = pd.DataFrame()
    df_C460 = pd.DataFrame()
    df_C465 = pd.DataFrame()
    df_C470 = pd.DataFrame()
    df_C480 = pd.DataFrame()
    df_C490 = pd.DataFrame()
    df_C495 = pd.DataFrame()
    df_C500 = pd.DataFrame()
    df_C510 = pd.DataFrame()
    df_C590 = pd.DataFrame()
    df_C591 = pd.DataFrame()
    df_C595 = pd.DataFrame()
    df_C597 = pd.DataFrame()
    df_C600 = pd.DataFrame()
    df_C601 = pd.DataFrame()
    df_C610 = pd.DataFrame()
    df_C690 = pd.DataFrame()
    df_C700 = pd.DataFrame()
    df_C790 = pd.DataFrame()
    df_C791 = pd.DataFrame()
    df_C800 = pd.DataFrame()
    df_C810 = pd.DataFrame()
    df_C815 = pd.DataFrame()
    df_C850 = pd.DataFrame()
    df_C855 = pd.DataFrame()
    df_C857 = pd.DataFrame()
    df_C860 = pd.DataFrame()
    df_C870 = pd.DataFrame()
    df_C880 = pd.DataFrame()
    df_C890 = pd.DataFrame()
    df_C895 = pd.DataFrame()
    df_C897 = pd.DataFrame()
    df_C990 = pd.DataFrame()
    # Bloco D
    df_D001 = pd.DataFrame()
    df_D100 = pd.DataFrame()
    df_D101 = pd.DataFrame()
    df_D110 = pd.DataFrame()
    df_D120 = pd.DataFrame()
    df_D130 = pd.DataFrame()
    df_D140 = pd.DataFrame()
    df_D150 = pd.DataFrame()
    df_D160 = pd.DataFrame()
    df_D161 = pd.DataFrame()
    df_D162 = pd.DataFrame()
    df_D170 = pd.DataFrame()
    df_D180 = pd.DataFrame()
    df_D190 = pd.DataFrame()
    df_D195 = pd.DataFrame()
    df_D197 = pd.DataFrame()
    df_D300 = pd.DataFrame()
    df_D301 = pd.DataFrame()
    df_D310 = pd.DataFrame()
    df_D350 = pd.DataFrame()
    df_D355 = pd.DataFrame()
    df_D360 = pd.DataFrame()
    df_D365 = pd.DataFrame()
    df_D370 = pd.DataFrame()
    df_D390 = pd.DataFrame()
    df_D400 = pd.DataFrame()
    df_D410 = pd.DataFrame()
    df_D411 = pd.DataFrame()
    df_D420 = pd.DataFrame()
    df_D500 = pd.DataFrame()
    df_D510 = pd.DataFrame()
    df_D530 = pd.DataFrame()
    df_D590 = pd.DataFrame()
    df_D600 = pd.DataFrame()
    df_D610 = pd.DataFrame()
    df_D690 = pd.DataFrame()
    df_D695 = pd.DataFrame()
    df_D696 = pd.DataFrame()
    df_D697 = pd.DataFrame()
    df_D700 = pd.DataFrame()
    df_D730 = pd.DataFrame()
    df_D731 = pd.DataFrame()
    df_D735 = pd.DataFrame()
    df_D737 = pd.DataFrame()
    df_D750 = pd.DataFrame()
    df_D760 = pd.DataFrame()
    df_D761 = pd.DataFrame()
    df_D990 = pd.DataFrame()
    # Bloco E
    df_E001 = pd.DataFrame()
    df_E100 = pd.DataFrame()
    df_E110 = pd.DataFrame()
    df_E111 = pd.DataFrame()
    df_E112 = pd.DataFrame()
    df_E113 = pd.DataFrame()
    df_E115 = pd.DataFrame()
    df_E116 = pd.DataFrame()
    df_E200 = pd.DataFrame()
    df_E210 = pd.DataFrame()
    df_E220 = pd.DataFrame()
    df_E230 = pd.DataFrame()
    df_E240 = pd.DataFrame()
    df_E250 = pd.DataFrame()
    df_E300 = pd.DataFrame()
    df_E310 = pd.DataFrame()
    df_E311 = pd.DataFrame()
    df_E312 = pd.DataFrame()
    df_E313 = pd.DataFrame()
    df_E316 = pd.DataFrame()
    df_E500 = pd.DataFrame()
    df_E510 = pd.DataFrame()
    df_E520 = pd.DataFrame()
    df_E530 = pd.DataFrame()
    df_E531 = pd.DataFrame()
    df_E990 = pd.DataFrame()
    # Bloco G
    df_G001 = pd.DataFrame()
    df_G110 = pd.DataFrame()
    df_G125 = pd.DataFrame()
    df_G126 = pd.DataFrame()
    df_G130 = pd.DataFrame()
    df_G140 = pd.DataFrame()
    df_G990 = pd.DataFrame()
    # Bloco H
    df_H001 = pd.DataFrame()
    df_H005 = pd.DataFrame()
    df_H010 = pd.DataFrame()
    df_H020 = pd.DataFrame()
    df_H030 = pd.DataFrame()
    df_H990 = pd.DataFrame()
    # Bloco K
    df_K001 = pd.DataFrame()
    df_K010 = pd.DataFrame()
    df_K100 = pd.DataFrame()
    df_K200 = pd.DataFrame()
    df_K210 = pd.DataFrame()
    df_K215 = pd.DataFrame()
    df_K220 = pd.DataFrame()
    df_K230 = pd.DataFrame()
    df_K235 = pd.DataFrame()
    df_K250 = pd.DataFrame()
    df_K255 = pd.DataFrame()
    df_K260 = pd.DataFrame()
    df_K265 = pd.DataFrame()
    df_K270 = pd.DataFrame()
    df_K275 = pd.DataFrame()
    df_K280 = pd.DataFrame()
    df_K290 = pd.DataFrame()
    df_K291 = pd.DataFrame()
    df_K292 = pd.DataFrame()
    df_K300 = pd.DataFrame()
    df_K301 = pd.DataFrame()
    df_K302 = pd.DataFrame()
    df_K990 = pd.DataFrame()
    # Bloco 1
    df_1001 = pd.DataFrame()
    df_1010 = pd.DataFrame()
    df_1100 = pd.DataFrame()
    df_1105 = pd.DataFrame()
    df_1110 = pd.DataFrame()
    df_1200 = pd.DataFrame()
    df_1210 = pd.DataFrame()
    df_1250 = pd.DataFrame()
    df_1255 = pd.DataFrame()
    df_1300 = pd.DataFrame()
    df_1310 = pd.DataFrame()
    df_1320 = pd.DataFrame()
    df_1350 = pd.DataFrame()
    df_1360 = pd.DataFrame()
    df_1370 = pd.DataFrame()
    df_1390 = pd.DataFrame()
    df_1391 = pd.DataFrame()
    df_1400 = pd.DataFrame()
    df_1500 = pd.DataFrame()
    df_1510 = pd.DataFrame()
    df_1600 = pd.DataFrame()
    df_1601 = pd.DataFrame()
    df_1700 = pd.DataFrame()
    df_1710 = pd.DataFrame()
    df_1800 = pd.DataFrame()
    df_1900 = pd.DataFrame()
    df_1910 = pd.DataFrame()
    df_1920 = pd.DataFrame()
    df_1921 = pd.DataFrame()
    df_1922 = pd.DataFrame()
    df_1923 = pd.DataFrame()
    df_1925 = pd.DataFrame()
    df_1926 = pd.DataFrame()
    df_1960 = pd.DataFrame()
    df_1970 = pd.DataFrame()
    df_1975 = pd.DataFrame()
    df_1980 = pd.DataFrame()
    df_1990 = pd.DataFrame()
    # Bloco 9
    df_9001 = pd.DataFrame()
    df_9900 = pd.DataFrame()
    df_9990 = pd.DataFrame()
    df_9999 = pd.DataFrame()

    df_0150 = pd.DataFrame()
    df_xxx = pd.DataFrame()

    # Função para atualizar a barra de progresso
    def update_progress(count):
        progress_bar["value"] = count
        root.update_idletasks()

    # Função para mostrar mensagem ao concluir
    def show_completion_message():
        showinfo("Concluído", "Processamento concluído com sucesso!")

    # Configurar janela
    root = Tk()
    root.title("Processando arquivos")
    root.geometry("300x100")

    # Criar barra de progresso
    progress_bar = ttk.Progressbar(root, length=200, mode="determinate")
    progress_bar.pack(pady=20)

    # Percorrer a lista de diretórios/arquivos em lotes com barra de progresso
    with tqdm(total=len(icms_list), desc="Processando arquivos", bar_format="{desc}") as pbar:
        for i in range(0, len(icms_list), batch_size):
            batch_paths = icms_list[i:i + batch_size]

            for path in batch_paths:
                try:
                    replace_icms_str = path.replace("/", "\\")

                    with open(replace_icms_str, 'r', encoding="latin-1") as file:
                        # Bloco 0
                        cont_0000 = {}
                        cont_0001 = {}
                        cont_0002 = {}
                        cont_0005 = {}
                        cont_0015 = {}
                        cont_0100 = {}
                        cont_0150 = {}
                        cont_0175 = {}
                        cont_0190 = {}
                        cont_0200 = {}
                        cont_0205 = {}
                        cont_0206 = {}
                        cont_0210 = {}
                        cont_0220 = {}
                        cont_0221 = {}
                        cont_0300 = {}
                        cont_0305 = {}
                        cont_0400 = {}
                        cont_0450 = {}
                        cont_0460 = {}
                        cont_0500 = {}
                        cont_0600 = {}
                        cont_0990 = {}
                        # Bloco B
                        cont_B001 = {}
                        cont_B020 = {}
                        cont_B025 = {}
                        cont_B030 = {}
                        cont_B035 = {}
                        cont_B350 = {}
                        cont_B420 = {}
                        cont_B440 = {}
                        cont_B460 = {}
                        cont_B470 = {}
                        cont_B500 = {}
                        cont_B510 = {}
                        cont_B990 = {}
                        # Bloco C
                        cont_C001 = {}
                        cont_C100 = {}
                        cont_C101 = {}
                        cont_C105 = {}
                        cont_C110 = {}
                        cont_C111 = {}
                        cont_C112 = {}
                        cont_C113 = {}
                        cont_C114 = {}
                        cont_C115 = {}
                        cont_C116 = {}
                        cont_C120 = {}
                        cont_C130 = {}
                        cont_C140 = {}
                        cont_C141 = {}
                        cont_C160 = {}
                        cont_C165 = {}
                        cont_C170 = {}
                        cont_C171 = {}
                        cont_C172 = {}
                        cont_C173 = {}
                        cont_C174 = {}
                        cont_C175 = {}
                        cont_C176 = {}
                        cont_C177 = {}
                        cont_C178 = {}
                        cont_C179 = {}
                        cont_C180 = {}
                        cont_C181 = {}
                        cont_C185 = {}
                        cont_C186 = {}
                        cont_C190 = {}
                        cont_C191 = {}
                        cont_C195 = {}
                        cont_C197 = {}
                        cont_C300 = {}
                        cont_C310 = {}
                        cont_C320 = {}
                        cont_C321 = {}
                        cont_C330 = {}
                        cont_C350 = {}
                        cont_C370 = {}
                        cont_C380 = {}
                        cont_C390 = {}
                        cont_C400 = {}
                        cont_C405 = {}
                        cont_C410 = {}
                        cont_C420 = {}
                        cont_C425 = {}
                        cont_C430 = {}
                        cont_C460 = {}
                        cont_C465 = {}
                        cont_C470 = {}
                        cont_C480 = {}
                        cont_C490 = {}
                        cont_C495 = {}
                        cont_C500 = {}
                        cont_C510 = {}
                        cont_C590 = {}
                        cont_C591 = {}
                        cont_C595 = {}
                        cont_C597 = {}
                        cont_C600 = {}
                        cont_C601 = {}
                        cont_C610 = {}
                        cont_C690 = {}
                        cont_C700 = {}
                        cont_C790 = {}
                        cont_C791 = {}
                        cont_C800 = {}
                        cont_C810 = {}
                        cont_C815 = {}
                        cont_C850 = {}
                        cont_C855 = {}
                        cont_C857 = {}
                        cont_C860 = {}
                        cont_C870 = {}
                        cont_C880 = {}
                        cont_C890 = {}
                        cont_C895 = {}
                        cont_C897 = {}
                        cont_C990 = {}
                        # Bloco D
                        cont_D001 = {}
                        cont_D100 = {}
                        cont_D101 = {}
                        cont_D110 = {}
                        cont_D120 = {}
                        cont_D130 = {}
                        cont_D140 = {}
                        cont_D150 = {}
                        cont_D160 = {}
                        cont_D161 = {}
                        cont_D162 = {}
                        cont_D170 = {}
                        cont_D180 = {}
                        cont_D190 = {}
                        cont_D195 = {}
                        cont_D197 = {}
                        cont_D300 = {}
                        cont_D301 = {}
                        cont_D310 = {}
                        cont_D350 = {}
                        cont_D355 = {}
                        cont_D360 = {}
                        cont_D365 = {}
                        cont_D370 = {}
                        cont_D390 = {}
                        cont_D400 = {}
                        cont_D410 = {}
                        cont_D411 = {}
                        cont_D420 = {}
                        cont_D500 = {}
                        cont_D510 = {}
                        cont_D530 = {}
                        cont_D590 = {}
                        cont_D600 = {}
                        cont_D610 = {}
                        cont_D690 = {}
                        cont_D695 = {}
                        cont_D696 = {}
                        cont_D697 = {}
                        cont_D700 = {}
                        cont_D730 = {}
                        cont_D731 = {}
                        cont_D735 = {}
                        cont_D737 = {}
                        cont_D750 = {}
                        cont_D760 = {}
                        cont_D761 = {}
                        cont_D990 = {}
                        # Bloco E
                        cont_E001 = {}
                        cont_E100 = {}
                        cont_E110 = {}
                        cont_E111 = {}
                        cont_E112 = {}
                        cont_E113 = {}
                        cont_E115 = {}
                        cont_E116 = {}
                        cont_E200 = {}
                        cont_E210 = {}
                        cont_E220 = {}
                        cont_E230 = {}
                        cont_E240 = {}
                        cont_E250 = {}
                        cont_E300 = {}
                        cont_E310 = {}
                        cont_E311 = {}
                        cont_E312 = {}
                        cont_E313 = {}
                        cont_E316 = {}
                        cont_E500 = {}
                        cont_E510 = {}
                        cont_E520 = {}
                        cont_E530 = {}
                        cont_E531 = {}
                        cont_E990 = {}
                        # Bloco G
                        cont_G001 = {}
                        cont_G110 = {}
                        cont_G125 = {}
                        cont_G126 = {}
                        cont_G130 = {}
                        cont_G140 = {}
                        cont_G990 = {}
                        # Bloco H
                        cont_H001 = {}
                        cont_H005 = {}
                        cont_H010 = {}
                        cont_H020 = {}
                        cont_H030 = {}
                        cont_H990 = {}
                        # Bloco K
                        cont_K001 = {}
                        cont_K010 = {}
                        cont_K100 = {}
                        cont_K200 = {}
                        cont_K210 = {}
                        cont_K215 = {}
                        cont_K220 = {}
                        cont_K230 = {}
                        cont_K235 = {}
                        cont_K250 = {}
                        cont_K255 = {}
                        cont_K260 = {}
                        cont_K265 = {}
                        cont_K270 = {}
                        cont_K275 = {}
                        cont_K280 = {}
                        cont_K290 = {}
                        cont_K291 = {}
                        cont_K292 = {}
                        cont_K300 = {}
                        cont_K301 = {}
                        cont_K302 = {}
                        cont_K990 = {}
                        # Bloco 1
                        cont_1001 = {}
                        cont_1010 = {}
                        cont_1100 = {}
                        cont_1105 = {}
                        cont_1110 = {}
                        cont_1200 = {}
                        cont_1210 = {}
                        cont_1250 = {}
                        cont_1255 = {}
                        cont_1300 = {}
                        cont_1310 = {}
                        cont_1320 = {}
                        cont_1350 = {}
                        cont_1360 = {}
                        cont_1370 = {}
                        cont_1390 = {}
                        cont_1391 = {}
                        cont_1400 = {}
                        cont_1500 = {}
                        cont_1510 = {}
                        cont_1600 = {}
                        cont_1601 = {}
                        cont_1700 = {}
                        cont_1710 = {}
                        cont_1800 = {}
                        cont_1900 = {}
                        cont_1910 = {}
                        cont_1920 = {}
                        cont_1921 = {}
                        cont_1922 = {}
                        cont_1923 = {}
                        cont_1925 = {}
                        cont_1926 = {}
                        cont_1960 = {}
                        cont_1970 = {}
                        cont_1975 = {}
                        cont_1980 = {}
                        cont_1990 = {}
                        # Bloco 9
                        cont_9001 = {}
                        cont_9900 = {}
                        cont_9990 = {}
                        cont_9999 = {}

                        for line in file:
                            first_six_chars = line[:6]

                            if '|0000|' in first_six_chars:
                                # Split the line by '|' and store the resulting columns in a list
                                columns_0000 = line.strip().split('|')
                                # assign to every column
                                # cont_0000['ID_DT_INI'] = data
                                # FormatarData(columns_0000, 7)
                                # cont_0000['ID_CNPJ'] = columns_0000.get(9, '') para teste caso for nulo futuramente
                                cont_0000['REG'] = columns_0000[1]
                                cont_0000['COD_VER'] = columns_0000[2]
                                cont_0000['COD_FIN'] = columns_0000[3]
                                cont_0000['DT_INI'] = columns_0000[4]
                                cont_0000['DT_FIN'] = columns_0000[5]
                                cont_0000['NOME'] = columns_0000[6]
                                cont_0000['CNPJ'] = columns_0000[7]
                                cont_0000['CPF'] = columns_0000[8]
                                cont_0000['UF'] = columns_0000[9]
                                cont_0000['IE'] = columns_0000[10]
                                cont_0000['COD_MUN'] = columns_0000[11]
                                cont_0000['IM'] = columns_0000[12]
                                cont_0000['SUFRAMA'] = columns_0000[13]
                                cont_0000['IND_PERFIL'] = columns_0000[14]
                                cont_0000['IND_ATIV'] = columns_0000[15]

                                # Create a new row in the dataframe using the columns list as the values
                                row_0000 = pd.DataFrame(cont_0000, index=[0])
                                # Concatenate the row to the dataframe
                                df_0000 = pd.concat([df_0000, row_0000])
                            elif '|0001|' in first_six_chars:
                                # Split the line by '|' and store the resulting columns in a list
                                columns_0001 = line.strip().split('|')
                                cont_0001['REG'] = columns_0001[1]
                                cont_0001['IND_MOV'] = columns_0001[2]
                                row_0001 = pd.DataFrame(cont_0001, index=[0])
                                df_0001 = pd.concat([df_0001, row_0001])
                            elif '|0002|' in first_six_chars:
                                columns_0002 = line.strip().split('|')
                                cont_0002['REG'] = columns_0002[1]
                                cont_0002['CLAS_ESTAB_IND'] = columns_0002[2]
                                row_0002 = pd.DataFrame(cont_0002, index=[0])
                                df_0002 = pd.concat([df_0002, row_0002])
                            elif '|0005|' in first_six_chars:
                                columns_0005 = line.strip().split('|')
                                cont_0005['REG'] = columns_0005[1]
                                cont_0005['FANTASIA'] = columns_0005[2]
                                cont_0005['CEP'] = columns_0005[3]
                                cont_0005['END'] = columns_0005[4]
                                cont_0005['NUM'] = columns_0005[5]
                                cont_0005['COMPL'] = columns_0005[6]
                                cont_0005['BAIRRO'] = columns_0005[7]
                                cont_0005['FONE'] = columns_0005[8]
                                cont_0005['FAX'] = columns_0005[9]
                                cont_0005['EMAIL'] = columns_0005[10]
                                row_0005 = pd.DataFrame(cont_0005, index=[0])
                                df_0005 = pd.concat([df_0005, row_0005])
                            elif '|0015|' in first_six_chars:
                                columns_0015 = line.strip().split('|')
                                cont_0015['REG'] = columns_0015[1]
                                cont_0015['UF_ST'] = columns_0015[2]
                                cont_0015['IE_ST'] = columns_0015[3]
                                row_0015 = pd.DataFrame(cont_0015, index=[0])
                                df_0015 = pd.concat([df_0015, row_0015])

                            elif '|0100|' in first_six_chars:
                                columns_0100 = line.strip().split('|')
                                cont_0100['REG'] = columns_0100[1]
                                cont_0100['NOME'] = columns_0100[2]
                                cont_0100['CPF'] = columns_0100[3]
                                cont_0100['CRC'] = columns_0100[4]
                                cont_0100['CNPJ'] = columns_0100[5]
                                cont_0100['CEP'] = columns_0100[6]
                                cont_0100['END'] = columns_0100[7]
                                cont_0100['NUM'] = columns_0100[8]
                                cont_0100['COMPL'] = columns_0100[9]
                                cont_0100['BAIRRO'] = columns_0100[10]
                                cont_0100['FONE'] = columns_0100[11]
                                cont_0100['FAX'] = columns_0100[12]
                                cont_0100['EMAIL'] = columns_0100[13]
                                cont_0100['COD_MUN'] = columns_0100[14]
                                row_0100 = pd.DataFrame(cont_0100, index=[0])
                                df_0100 = pd.concat([df_0100, row_0100])
                            elif '|0150|' in first_six_chars:
                                columns_0150 = line.strip().split('|')
                                cont_0150['REG'] = columns_0150[1]
                                cont_0150['COD_PART'] = columns_0150[2]
                                cont_0150['NOME'] = columns_0150[3]
                                cont_0150['COD_PAIS'] = columns_0150[4]
                                cont_0150['CNPJ'] = columns_0150[5]
                                cont_0150['CPF'] = columns_0150[6]
                                cont_0150['IE'] = columns_0150[7]
                                cont_0150['COD_MUN'] = columns_0150[8]
                                cont_0150['SUFRAMA'] = columns_0150[9]
                                cont_0150['END'] = columns_0150[10]
                                cont_0150['NUM'] = columns_0150[11]
                                cont_0150['COMPL'] = columns_0150[12]
                                cont_0150['BAIRRO'] = columns_0150[13]
                                row_0150 = pd.DataFrame(cont_0150, index=[0])
                                df_0150 = pd.concat([df_0150, row_0150])
                            elif '|0175|' in first_six_chars:
                                columns_0175 = line.strip().split('|')
                                cont_0175['REG'] = columns_0175[1]
                                cont_0175['DT_ALT'] = columns_0175[2]
                                cont_0175['NR_CAMPO'] = columns_0175[3]
                                cont_0175['CONT_ANT'] = columns_0175[4]
                                row_0175 = pd.DataFrame(cont_0175, index=[0])
                                df_0175 = pd.concat([df_0175, row_0175])
                            elif '|0190|' in first_six_chars:
                                columns_0190 = line.strip().split('|')
                                cont_0190['REG'] = columns_0190[1]
                                cont_0190['UNID'] = columns_0190[2]
                                cont_0190['DESCR'] = columns_0190[3]
                                row_0190 = pd.DataFrame(cont_0190, index=[0])
                                df_0190 = pd.concat([df_0190, row_0190])
                            elif '|0200|' in first_six_chars:
                                columns_0200 = line.strip().split('|')
                                cont_0200['REG'] = columns_0200[1]
                                cont_0200['COD_ITEM'] = columns_0200[2]
                                cont_0200['DESCR_ITEM'] = columns_0200[3]
                                cont_0200['COD_BARRA'] = columns_0200[4]
                                cont_0200['COD_ANT_ITEM'] = columns_0200[5]
                                cont_0200['UNID_INV'] = columns_0200[6]
                                cont_0200['TIPO_ITEM'] = columns_0200[7]
                                cont_0200['COD_NCM'] = columns_0200[8]
                                cont_0200['EX_IPI'] = columns_0200[9]
                                cont_0200['COD_GEN'] = columns_0200[10]
                                cont_0200['COD_LST'] = columns_0200[11]
                                cont_0200['ALIQ_ICMS'] = columns_0200[12]
                                cont_0200['CEST'] = columns_0200[13]
                                row_0200 = pd.DataFrame(cont_0200, index=[0])
                                df_0200 = pd.concat([df_0200, row_0200])

                            elif '|0205|' in first_six_chars:
                                columns_0205 = line.strip().split('|')
                                cont_0205['REG'] = columns_0205[1]
                                cont_0205['DESCR_ANT_ITEM'] = columns_0205[2]
                                cont_0205['DT_INI'] = columns_0205[3]
                                cont_0205['DT_FIM'] = columns_0205[4]
                                cont_0205['COD_ANT_ITEM'] = columns_0205[5]
                                row_0205 = pd.DataFrame(cont_0205, index=[0])
                                df_0205 = pd.concat([df_0205, row_0205])

                            elif '|0206|' in first_six_chars:
                                columns_0206 = line.strip().split('|')
                                cont_0206['REG'] = columns_0206[1]
                                cont_0206['COD_COMB'] = columns_0206[2]
                                row_0206 = pd.DataFrame(cont_0206, index=[0])
                                df_0206 = pd.concat([df_0206, row_0206])
                            elif '|0210|' in first_six_chars:
                                columns_0210 = line.strip().split('|')
                                cont_0210['REG'] = columns_0210[1]
                                cont_0210['COD_ITEM_COMP'] = columns_0210[2]
                                cont_0210['QTD_COMP'] = columns_0210[3]
                                cont_0210['PERDA'] = columns_0210[4]
                                row_0210 = pd.DataFrame(cont_0210, index=[0])
                                df_0210 = pd.concat([df_0210, row_0210])
                            elif '|0220|' in first_six_chars:
                                columns_0220 = line.strip().split('|')
                                cont_0220['REG'] = columns_0220[1]
                                cont_0220['UNID_CONV'] = columns_0220[2]
                                cont_0220['FAT_CONV'] = columns_0220[3]
                                cont_0220['COD_BARRA'] = columns_0220[4]
                                row_0220 = pd.DataFrame(cont_0220, index=[0])
                                df_0220 = pd.concat([df_0220, row_0220])
                            elif '|0221|' in first_six_chars:
                                columns_0221 = line.strip().split('|')
                                cont_0221['REG'] = columns_0221[1]
                                cont_0221['UNID_CONV'] = columns_0221[2]
                                cont_0221['FAT_CONV'] = columns_0221[3]
                                row_0221 = pd.DataFrame(cont_0221, index=[0])
                                df_0221 = pd.concat([df_0221, row_0221])
                            elif '|0300|' in first_six_chars:
                                columns_0300 = line.strip().split('|')
                                cont_0300['REG'] = columns_0300[1]
                                cont_0300['COD_IND_BEM'] = columns_0300[2]
                                cont_0300['IDENT_MERC'] = columns_0300[3]
                                cont_0300['DESCR_ITEM'] = columns_0300[4]
                                cont_0300['COD_PRNC'] = columns_0300[5]
                                cont_0300['COD_CTA'] = columns_0300[6]
                                cont_0300['NR_PARC'] = columns_0300[7]
                                row_0300 = pd.DataFrame(cont_0300, index=[0])
                                df_0300 = pd.concat([df_0300, row_0300])
                            elif '|0305|' in first_six_chars:
                                columns_0305 = line.strip().split('|')
                                cont_0305['REG'] = columns_0305[1]
                                cont_0305['COD_CCUS'] = columns_0305[2]
                                cont_0305['FUNC'] = columns_0305[3]
                                cont_0305['VIDA_UTIL'] = columns_0305[4]
                                row_0305 = pd.DataFrame(cont_0305, index=[0])
                                df_0305 = pd.concat([df_0305, row_0305])
                            elif '|0400|' in first_six_chars:
                                columns_0400 = line.strip().split('|')
                                cont_0400['REG'] = columns_0400[1]
                                cont_0400['COD_NAT'] = columns_0400[2]
                                cont_0400['DESCR_NAT'] = columns_0400[3]
                                row_0400 = pd.DataFrame(cont_0400, index=[0])
                                df_0400 = pd.concat([df_0400, row_0400])
                            elif '|0450|' in first_six_chars:
                                columns_0450 = line.strip().split('|')
                                cont_0450['REG'] = columns_0450[1]
                                cont_0450['COD_INF'] = columns_0450[2]
                                cont_0450['TXT'] = columns_0450[3]
                                row_0450 = pd.DataFrame(cont_0450, index=[0])
                                df_0450 = pd.concat([df_0450, row_0450])
                            elif '|0460|' in first_six_chars:
                                columns_0460 = line.strip().split('|')
                                cont_0460['REG'] = columns_0460[1]
                                cont_0460['COD_OBS'] = columns_0460[2]
                                cont_0460['TXT'] = columns_0460[3]
                                row_0460 = pd.DataFrame(cont_0460, index=[0])
                                df_0460 = pd.concat([df_0460, row_0460])
                            elif '|0500|' in first_six_chars:
                                columns_0500 = line.strip().split('|')
                                cont_0500['REG'] = columns_0500[1]
                                cont_0500['DT_ALT'] = columns_0500[2]
                                cont_0500['COD_NAT_CC'] = columns_0500[3]
                                cont_0500['IND_CTA'] = columns_0500[4]
                                cont_0500['NÍVEL'] = columns_0500[5]
                                cont_0500['COD_CTA'] = columns_0500[6]
                                cont_0500['NOME_CTA'] = columns_0500[7]
                                row_0500 = pd.DataFrame(cont_0500, index=[0])
                                df_0500 = pd.concat([df_0500, row_0500])
                            elif '|0600|' in first_six_chars:
                                columns_0600 = line.strip().split('|')
                                cont_0600['REG'] = columns_0600[1]
                                cont_0600['DT_ALT'] = columns_0600[2]
                                cont_0600['COD_CCUS'] = columns_0600[3]
                                cont_0600['CCUS'] = columns_0600[4]
                                row_0600 = pd.DataFrame(cont_0600, index=[0])
                                df_0600 = pd.concat([df_0600, row_0600])
                            elif '|0990|' in first_six_chars:
                                columns_0990 = line.strip().split('|')
                                cont_0990['REG'] = columns_0990[1]
                                cont_0990['QTD_LIN_0'] = columns_0990[2]
                                row_0990 = pd.DataFrame(cont_0990, index=[0])
                                df_0990 = pd.concat([df_0990, row_0990])
                            elif '|B001|' in first_six_chars:
                                columns_B001 = line.strip().split('|')
                                cont_B001['REG'] = columns_B001[1]
                                cont_B001['IND_DAD'] = columns_B001[2]
                                row_B001 = pd.DataFrame(cont_B001, index=[0])
                                df_B001 = pd.concat([df_B001, row_B001])
                            elif '|B020|' in first_six_chars:
                                columns_B020 = line.strip().split('|')
                                cont_B020['REG'] = columns_B020[1]
                                cont_B020['IND_OPER'] = columns_B020[2]
                                cont_B020['IND_EMIT'] = columns_B020[3]
                                cont_B020['COD_PART'] = columns_B020[4]
                                cont_B020['COD_MOD'] = columns_B020[5]
                                cont_B020['COD_SIT'] = columns_B020[6]
                                cont_B020['SER'] = columns_B020[7]
                                cont_B020['NUM_DOC'] = columns_B020[8]
                                cont_B020['CHV_NFE'] = columns_B020[9]
                                cont_B020['DT_DOC'] = columns_B020[10]
                                cont_B020['COD_MUN_SE_RV'] = columns_B020[11]
                                cont_B020['VL_CONT'] = columns_B020[12]
                                cont_B020['VL_MAT_TERC'] = columns_B020[13]
                                cont_B020['VL_SUB'] = columns_B020[14]
                                cont_B020['VL_ISNT_ISS'] = columns_B020[15]
                                cont_B020['VL_DED_BC'] = columns_B020[16]
                                cont_B020['VL_BC_ISS'] = columns_B020[17]
                                cont_B020['VL_BC_ISS_RT'] = columns_B020[18]
                                cont_B020['VL_ISS_RT'] = columns_B020[19]
                                cont_B020['VL_ISS'] = columns_B020[20]
                                cont_B020['COD_INF_OBS'] = columns_B020[21]
                                row_B020 = pd.DataFrame(cont_B020, index=[0])
                                df_B020 = pd.concat([df_B020, row_B020])
                            elif '|B025|' in first_six_chars:
                                columns_B025 = line.strip().split('|')
                                cont_B025['REG'] = columns_B025[1]
                                cont_B025['VL_CONT_P'] = columns_B025[2]
                                cont_B025['VL_BC_ISS_P'] = columns_B025[3]
                                cont_B025['ALIQ_ISS'] = columns_B025[4]
                                cont_B025['VL_ISS_P'] = columns_B025[5]
                                cont_B025['VL_ISNT_ISS_P'] = columns_B025[6]
                                cont_B025['COD_SERV'] = columns_B025[7]
                                row_B025 = pd.DataFrame(cont_B025, index=[0])
                                df_B025 = pd.concat([df_B025, row_B025])
                            elif '|B030|' in first_six_chars:
                                columns_B030 = line.strip().split('|')
                                cont_B030['REG'] = columns_B030[1]
                                cont_B030['COD_MOD'] = columns_B030[2]
                                cont_B030['SER'] = columns_B030[3]
                                cont_B030['NUM_DOC_INI'] = columns_B030[4]
                                cont_B030['NUM_DOC_FIN'] = columns_B030[5]
                                cont_B030['DT_DOC'] = columns_B030[6]
                                cont_B030['QTD_CANC'] = columns_B030[7]
                                cont_B030['VL_CONT'] = columns_B030[8]
                                cont_B030['VL_ISNT_ISS'] = columns_B030[9]
                                cont_B030['VL_BC_ISS'] = columns_B030[10]
                                cont_B030['VL_ISS'] = columns_B030[11]
                                cont_B030['COD_INF_OBS'] = columns_B030[12]
                                row_B030 = pd.DataFrame(cont_B030, index=[0])
                                df_B030 = pd.concat([df_B030, row_B030])
                            elif '|B035|' in first_six_chars:
                                columns_B035 = line.strip().split('|')
                                cont_B035['REG'] = columns_B035[1]
                                cont_B035['VL_CONT_P'] = columns_B035[2]
                                cont_B035['VL_BC_ISS_P'] = columns_B035[3]
                                cont_B035['ALIQ_ISS'] = columns_B035[4]
                                cont_B035['VL_ISS_P'] = columns_B035[5]
                                cont_B035['VL_ISNT_ISS_P'] = columns_B035[6]
                                cont_B035['COD_SERV'] = columns_B035[7]
                                row_B035 = pd.DataFrame(cont_B035, index=[0])
                                df_B035 = pd.concat([df_B035, row_B035])
                            elif '|B350|' in first_six_chars:
                                columns_B350 = line.strip().split('|')
                                cont_B350['REG'] = columns_B350[1]
                                cont_B350['COD_CTD'] = columns_B350[2]
                                cont_B350['CTA_ISS'] = columns_B350[3]
                                cont_B350['CTA_COSIF'] = columns_B350[4]
                                cont_B350['QTD_OCOR'] = columns_B350[5]
                                cont_B350['COD_SERV'] = columns_B350[6]
                                cont_B350['VL_CONT'] = columns_B350[7]
                                cont_B350['VL_BC_ISS'] = columns_B350[8]
                                cont_B350['ALIQ_ISS'] = columns_B350[9]
                                cont_B350['VL_ISS'] = columns_B350[10]
                                cont_B350['COD_INF_OBS'] = columns_B350[11]
                                row_B350 = pd.DataFrame(cont_B350, index=[0])
                                df_B350 = pd.concat([df_B350, row_B350])
                            elif '|B420|' in first_six_chars:
                                columns_B420 = line.strip().split('|')
                                cont_B420['REG'] = columns_B420[1]
                                cont_B420['VL_CONT'] = columns_B420[2]
                                cont_B420['VL_BC_ISS'] = columns_B420[3]
                                cont_B420['ALIQ_ISS'] = columns_B420[4]
                                cont_B420['VL_ISNT_ISS'] = columns_B420[5]
                                cont_B420['VL_ISS'] = columns_B420[6]
                                cont_B420['COD_SERV'] = columns_B420[7]
                                row_B420 = pd.DataFrame(cont_B420, index=[0])
                                df_B420 = pd.concat([df_B420, row_B420])
                            elif '|B440|' in first_six_chars:
                                columns_B440 = line.strip().split('|')
                                cont_B440['REG'] = columns_B440[1]
                                cont_B440['IND_OPER'] = columns_B440[2]
                                cont_B440['COD_PART'] = columns_B440[3]
                                cont_B440['VL_CONT_RT'] = columns_B440[4]
                                cont_B440['VL_BC_ISS_RT'] = columns_B440[5]
                                cont_B440['VL_ISS_RT'] = columns_B440[6]
                                row_B440 = pd.DataFrame(cont_B440, index=[0])
                                df_B440 = pd.concat([df_B440, row_B440])
                            elif '|B460|' in first_six_chars:
                                columns_B460 = line.strip().split('|')
                                cont_B460['REG'] = columns_B460[1]
                                cont_B460['IND_DED'] = columns_B460[2]
                                cont_B460['VL_DED'] = columns_B460[3]
                                cont_B460['NUM_PROC'] = columns_B460[4]
                                cont_B460['IND_PROC'] = columns_B460[5]
                                cont_B460['PROC'] = columns_B460[6]
                                cont_B460['COD_INF_OBS'] = columns_B460[7]
                                cont_B460['IND_OBR'] = columns_B460[8]
                                row_B460 = pd.DataFrame(cont_B460, index=[0])
                                df_B460 = pd.concat([df_B460, row_B460])
                            elif '|B470|' in first_six_chars:
                                columns_B470 = line.strip().split('|')
                                cont_B470['REG'] = columns_B470[1]
                                cont_B470['VL_CONT'] = columns_B470[2]
                                cont_B470['VL_MAT_TERC'] = columns_B470[3]
                                cont_B470['VL_MAT_PROP'] = columns_B470[4]
                                cont_B470['VL_SUB'] = columns_B470[5]
                                cont_B470['VL_ISNT'] = columns_B470[6]
                                cont_B470['VL_DED_BC'] = columns_B470[7]
                                cont_B470['VL_BC_ISS'] = columns_B470[8]
                                cont_B470['VL_BC_ISS_RT'] = columns_B470[9]
                                cont_B470['VL_ISS'] = columns_B470[10]
                                cont_B470['VL_ISS_RT'] = columns_B470[11]
                                cont_B470['VL_DED'] = columns_B470[12]
                                cont_B470['VL_ISS_REC'] = columns_B470[13]
                                cont_B470['VL_ISS_ST'] = columns_B470[14]
                                cont_B470['VL_ISS_REC_UNI'] = columns_B470[15]
                                row_B470 = pd.DataFrame(cont_B470, index=[0])
                                df_B470 = pd.concat([df_B470, row_B470])
                            elif '|B500|' in first_six_chars:
                                columns_B500 = line.strip().split('|')
                                cont_B500['REG'] = columns_B500[1]
                                cont_B500['VL_REC'] = columns_B500[2]
                                cont_B500['QTD_PROF'] = columns_B500[3]
                                cont_B500['VL_OR'] = columns_B500[4]
                                row_B500 = pd.DataFrame(cont_B500, index=[0])
                                df_B500 = pd.concat([df_B500, row_B500])
                            elif '|B510|' in first_six_chars:
                                columns_B510 = line.strip().split('|')
                                cont_B510['REG'] = columns_B510[1]
                                cont_B510['IND_PROF'] = columns_B510[2]
                                cont_B510['IND_ESC'] = columns_B510[3]
                                cont_B510['IND_SOC'] = columns_B510[4]
                                cont_B510['CPF'] = columns_B510[5]
                                cont_B510['NOME'] = columns_B510[6]
                                row_B510 = pd.DataFrame(cont_B510, index=[0])
                                df_B510 = pd.concat([df_B510, row_B510])
                            elif '|B990|' in first_six_chars:
                                columns_B990 = line.strip().split('|')
                                cont_B990['REG'] = columns_B990[1]
                                cont_B990['QTD_LIN_B'] = columns_B990[2]
                                row_B990 = pd.DataFrame(cont_B990, index=[0])
                                df_B990 = pd.concat([df_B990, row_B990])
                            elif '|C001|' in first_six_chars:
                                columns_C001 = line.strip().split('|')
                                cont_C001['REG'] = columns_C001[1]
                                cont_C001['IND_MOV'] = columns_C001[2]
                                row_C001 = pd.DataFrame(cont_C001, index=[0])
                                df_C001 = pd.concat([df_C001, row_C001])
                            elif '|C100|' in first_six_chars:
                                columns_C100 = line.strip().split('|')
                                cont_C100['REG'] = columns_C100[1]
                                cont_C100['IND_OPER'] = columns_C100[2]
                                cont_C100['IND_EMIT'] = columns_C100[3]
                                cont_C100['COD_PART'] = columns_C100[4]
                                cont_C100['COD_MOD'] = columns_C100[5]
                                cont_C100['COD_SIT'] = columns_C100[6]
                                cont_C100['SER'] = columns_C100[7]
                                cont_C100['NUM_DOC'] = columns_C100[8]
                                cont_C100['CHV_NFE'] = columns_C100[9]
                                cont_C100['DT_DOC'] = columns_C100[10]
                                cont_C100['DT_E_S'] = columns_C100[11]
                                cont_C100['VL_DOC'] = columns_C100[12]
                                cont_C100['IND_PGTO'] = columns_C100[13]
                                cont_C100['VL_DESC'] = columns_C100[14]
                                cont_C100['VL_ABAT_NT'] = columns_C100[15]
                                cont_C100['VL_MERC'] = columns_C100[16]
                                cont_C100['IND_FRT'] = columns_C100[17]
                                cont_C100['VL_FRT'] = columns_C100[18]
                                cont_C100['VL_SEG'] = columns_C100[19]
                                cont_C100['VL_OUT_DA'] = columns_C100[20]
                                cont_C100['VL_BC_ICMS'] = columns_C100[21]
                                cont_C100['VL_ICMS'] = columns_C100[22]
                                cont_C100['VL_BC_ICMS_ST'] = columns_C100[23]
                                cont_C100['VL_ICMS_ST'] = columns_C100[24]
                                cont_C100['VL_IPI'] = columns_C100[25]
                                cont_C100['VL_PIS'] = columns_C100[26]
                                cont_C100['VL_COFINS'] = columns_C100[27]
                                cont_C100['VL_PIS_ST'] = columns_C100[28]
                                cont_C100['VL_COFINS_ST'] = columns_C100[29]
                                row_C100 = pd.DataFrame(cont_C100, index=[0])
                                df_C100 = pd.concat([df_C100, row_C100])
                            elif '|C101|' in first_six_chars:
                                columns_C101 = line.strip().split('|')
                                cont_C101['REG'] = columns_C101[1]
                                cont_C101['VL_FCP_UF_DEST'] = columns_C101[2]
                                cont_C101['VL_ICMS_UF_DEST'] = columns_C101[3]
                                cont_C101['VL_ICMS_UF_REM'] = columns_C101[4]
                                row_C101 = pd.DataFrame(cont_C101, index=[0])
                                df_C101 = pd.concat([df_C101, row_C101])
                            elif '|C105|' in first_six_chars:
                                columns_C105 = line.strip().split('|')
                                cont_C105['REG'] = columns_C105[1]
                                cont_C105['OPER'] = columns_C105[2]
                                cont_C105['UF'] = columns_C105[3]
                                row_C105 = pd.DataFrame(cont_C105, index=[0])
                                df_C105 = pd.concat([df_C105, row_C105])
                            elif '|C110|' in first_six_chars:
                                columns_C110 = line.strip().split('|')
                                cont_C110['REG'] = columns_C110[1]
                                cont_C110['COD_INF'] = columns_C110[2]
                                cont_C110['TXT_COMPL'] = columns_C110[3]
                                row_C110 = pd.DataFrame(cont_C110, index=[0])
                                df_C110 = pd.concat([df_C110, row_C110])
                            elif '|C111|' in first_six_chars:
                                columns_C111 = line.strip().split('|')
                                cont_C111['REG'] = columns_C111[1]
                                cont_C111['NUM_PROC'] = columns_C111[2]
                                cont_C111['IND_PROC'] = columns_C111[3]
                                row_C111 = pd.DataFrame(cont_C111, index=[0])
                                df_C111 = pd.concat([df_C111, row_C111])
                            elif '|C112|' in first_six_chars:
                                columns_C112 = line.strip().split('|')
                                cont_C112['REG'] = columns_C112[1]
                                cont_C112['COD_DA'] = columns_C112[2]
                                cont_C112['UF'] = columns_C112[3]
                                cont_C112['NUM_DA'] = columns_C112[4]
                                cont_C112['COD_AUT'] = columns_C112[5]
                                cont_C112['VL_DA'] = columns_C112[6]
                                cont_C112['DT_VCTO'] = columns_C112[7]
                                cont_C112['DT_PGTO'] = columns_C112[8]
                                row_C112 = pd.DataFrame(cont_C112, index=[0])
                                df_C112 = pd.concat([df_C112, row_C112])
                            elif '|C113|' in first_six_chars:
                                columns_C113 = line.strip().split('|')
                                cont_C113['REG'] = columns_C113[1]
                                cont_C113['IND_OPER'] = columns_C113[2]
                                cont_C113['IND_EMIT'] = columns_C113[3]
                                cont_C113['COD_PART'] = columns_C113[4]
                                cont_C113['COD_MOD'] = columns_C113[5]
                                cont_C113['SER'] = columns_C113[6]
                                cont_C113['SUB'] = columns_C113[7]
                                cont_C113['NUM_DOC'] = columns_C113[8]
                                cont_C113['DT_DOC'] = columns_C113[9]
                                cont_C113['CHV_DOCe'] = columns_C113[10]
                                row_C113 = pd.DataFrame(cont_C113, index=[0])
                                df_C113 = pd.concat([df_C113, row_C113])
                            elif '|C114|' in first_six_chars:
                                columns_C114 = line.strip().split('|')
                                cont_C114['REG'] = columns_C114[1]
                                cont_C114['COD_MOD'] = columns_C114[2]
                                cont_C114['ECF_FAB'] = columns_C114[3]
                                cont_C114['ECF_CX'] = columns_C114[4]
                                cont_C114['NUM_DOC'] = columns_C114[5]
                                cont_C114['DT_DOC'] = columns_C114[6]
                                row_C114 = pd.DataFrame(cont_C114, index=[0])
                                df_C114 = pd.concat([df_C114, row_C114])
                            elif '|C115|' in first_six_chars:
                                columns_C115 = line.strip().split('|')
                                cont_C115['REG'] = columns_C115[1]
                                cont_C115['IND_CARGA'] = columns_C115[2]
                                cont_C115['CNPJ_COL'] = columns_C115[3]
                                cont_C115['IE_COL'] = columns_C115[4]
                                cont_C115['CPF_COL'] = columns_C115[5]
                                cont_C115['COD_MUN_COL'] = columns_C115[6]
                                cont_C115['CNPJ_ENTG'] = columns_C115[7]
                                cont_C115['IE_ENTG'] = columns_C115[8]
                                cont_C115['CPF_ENTG'] = columns_C115[9]
                                cont_C115['COD_MUN_ENTG'] = columns_C115[10]#######################################DUVIDA
                                row_C115 = pd.DataFrame(cont_C115, index=[0])
                                df_C115 = pd.concat([df_C115, row_C115])
                            elif '|C116|' in first_six_chars:
                                columns_C116 = line.strip().split('|')
                                cont_C116['REG'] = columns_C116[1]
                                cont_C116['COD_MOD'] = columns_C116[2]
                                cont_C116['NR_SAT'] = columns_C116[3]
                                cont_C116['CHV_CFE'] = columns_C116[4]
                                cont_C116['NUM_CFE'] = columns_C116[5]
                                cont_C116['DT_DOC'] = columns_C116[6]
                                row_C116 = pd.DataFrame(cont_C116, index=[0])
                                df_C116 = pd.concat([df_C116, row_C116])
                            elif '|C120|' in first_six_chars:
                                columns_C120 = line.strip().split('|')
                                cont_C120['REG'] = columns_C120[1]
                                cont_C120['COD_DOC_IMP'] = columns_C120[2]
                                cont_C120['NUM_DOC_IMP'] = columns_C120[3]
                                cont_C120['PIS_IMP'] = columns_C120[4]
                                cont_C120['COFINS_IMP'] = columns_C120[5]####################################
                                cont_C120['NUM_ACDRAW'] = columns_C120[6]
                                row_C120 = pd.DataFrame(cont_C120, index=[0])
                                df_C120 = pd.concat([df_C120, row_C120])
                            elif '|C130|' in first_six_chars:
                                columns_C130 = line.strip().split('|')
                                cont_C130['REG'] = columns_C130[1]
                                cont_C130['VL_SERV_NT'] = columns_C130[2]
                                cont_C130['VL_BC_ISSQN'] = columns_C130[3]
                                cont_C130['VL_ISSQN'] = columns_C130[4]
                                cont_C130['VL_BC_IRRF'] = columns_C130[5]
                                cont_C130['VL_IRRF'] = columns_C130[6]
                                cont_C130['VL_BC_PREV'] = columns_C130[7]
                                cont_C130['VL_PREV'] = columns_C130[8]
                                row_C130 = pd.DataFrame(cont_C130, index=[0])
                                df_C130 = pd.concat([df_C130, row_C130])
                            elif '|C140|' in first_six_chars:
                                columns_C140 = line.strip().split('|')
                                cont_C140['REG'] = columns_C140[1]
                                cont_C140['IND_EMIT'] = columns_C140[2]
                                cont_C140['IND_TIT'] = columns_C140[3]
                                cont_C140['DESC_TIT'] = columns_C140[4]
                                cont_C140['NUM_TIT'] = columns_C140[5]
                                cont_C140['QTD_PARC'] = columns_C140[6]
                                cont_C140['VL_TIT'] = columns_C140[7]
                                row_C140 = pd.DataFrame(cont_C140, index=[0])
                                df_C140 = pd.concat([df_C140, row_C140])
                            elif '|C141|' in first_six_chars:
                                columns_C141 = line.strip().split('|')
                                cont_C141['REG'] = columns_C141[1]
                                cont_C141['NUM_PARC'] = columns_C141[2]
                                cont_C141['DT_VCTO'] = columns_C141[3]
                                cont_C141['VL_PARC'] = columns_C141[4]
                                row_C141 = pd.DataFrame(cont_C141, index=[0])
                                df_C141 = pd.concat([df_C141, row_C141])
                            elif '|C160|' in first_six_chars:
                                columns_C160 = line.strip().split('|')
                                cont_C160['REG'] = columns_C160[1]
                                cont_C160['COD_PART'] = columns_C160[2]
                                cont_C160['VEIC_ID'] = columns_C160[3]
                                cont_C160['QTD_VOL'] = columns_C160[4]
                                cont_C160['PESO_BRT'] = columns_C160[5]
                                cont_C160['PESO_LIQ'] = columns_C160[6]
                                cont_C160['UF_ID'] = columns_C160[7]
                                row_C160 = pd.DataFrame(cont_C160, index=[0])
                                df_C160 = pd.concat([df_C160, row_C160])
                            elif '|C165|' in first_six_chars:
                                columns_C165 = line.strip().split('|')
                                cont_C165['REG'] = columns_C165[1]
                                cont_C165['COD_PART'] = columns_C165[2]
                                cont_C165['VEIC_ID'] = columns_C165[3]
                                cont_C165['COD_AUT'] = columns_C165[4]
                                cont_C165['NR_PASSE'] = columns_C165[5]
                                cont_C165['HORA'] = columns_C165[6]
                                cont_C165['TEMPER'] = columns_C165[7]
                                cont_C165['QTD_VOL'] = columns_C165[8]
                                cont_C165['PESO_BRT'] = columns_C165[9]
                                cont_C165['PESO_LIQ'] = columns_C165[10]
                                cont_C165['NOM_MOT'] = columns_C165[11]
                                cont_C165['CPF'] = columns_C165[12]
                                cont_C165['UF_ID'] = columns_C165[13]
                                row_C165 = pd.DataFrame(cont_C165, index=[0])
                                df_C165 = pd.concat([df_C165, row_C165])
                            elif '|C170|' in first_six_chars:
                                columns_C170 = line.strip().split('|')
                                cont_C170['REG'] = columns_C170[1]
                                cont_C170['NUM_ITEM'] = columns_C170[2]
                                cont_C170['COD_ITEM'] = columns_C170[3]
                                cont_C170['DESCR_COMPL'] = columns_C170[4]
                                cont_C170['QTD'] = columns_C170[5]
                                cont_C170['UNID'] = columns_C170[6]
                                cont_C170['VL_ITEM'] = columns_C170[7]
                                cont_C170['VL_DESC'] = columns_C170[8]
                                cont_C170['IND_MOV'] = columns_C170[9]
                                cont_C170['CST_ICMS'] = columns_C170[10]
                                cont_C170['CFOP'] = columns_C170[11]
                                cont_C170['COD_NAT'] = columns_C170[12]
                                cont_C170['VL_BC_ICMS'] = columns_C170[13]
                                cont_C170['ALIQ_ICMS'] = columns_C170[14]
                                cont_C170['VL_ICMS'] = columns_C170[15]
                                cont_C170['VL_BC_ICMS_ST'] = columns_C170[16]
                                cont_C170['ALIQ_ST'] = columns_C170[17]
                                cont_C170['VL_ICMS_ST'] = columns_C170[18]
                                cont_C170['IND_APUR'] = columns_C170[19]
                                cont_C170['CST_IPI'] = columns_C170[20]
                                cont_C170['COD_ENQ'] = columns_C170[21]
                                cont_C170['VL_BC_IPI'] = columns_C170[22]
                                cont_C170['ALIQ_IPI'] = columns_C170[23]
                                cont_C170['VL_IPI'] = columns_C170[24]
                                cont_C170['CST_PIS'] = columns_C170[25]
                                cont_C170['VL_BC_PIS'] = columns_C170[26]
                                cont_C170['ALIQ_PIS'] = columns_C170[27]
                                cont_C170['QUANT_BC_PIS'] = columns_C170[28]
                                cont_C170['ALIQ_PIS'] = columns_C170[29]
                                cont_C170['VL_PIS'] = columns_C170[30]
                                cont_C170['CST_COFINS'] = columns_C170[31]
                                cont_C170['VL_BC_COFINS'] = columns_C170[32]
                                cont_C170['ALIQ_COFINS'] = columns_C170[33]
                                cont_C170['QUANT_BC_COF_INS'] = columns_C170[34]################################
                                cont_C170['ALIQ_COFINS'] = columns_C170[35]
                                cont_C170['VL_COFINS'] = columns_C170[36]
                                cont_C170['COD_CTA'] = columns_C170[37]
                                cont_C170['VL_ABAT_NT'] = columns_C170[38]
                                row_C170 = pd.DataFrame(cont_C170, index=[0])
                                df_C170 = pd.concat([df_C170, row_C170])
                            elif '|C171|' in first_six_chars:
                                columns_C171 = line.strip().split('|')
                                cont_C171['REG'] = columns_C171[1]
                                cont_C171['NUM_TANQUE'] = columns_C171[2]
                                cont_C171['QTDE'] = columns_C171[3]
                                row_C171 = pd.DataFrame(cont_C171, index=[0])
                                df_C171 = pd.concat([df_C171, row_C171])
                            elif '|C172|' in first_six_chars:
                                columns_C172 = line.strip().split('|')
                                cont_C172['REG'] = columns_C172[1]
                                cont_C172['VL_BC_ISSQN'] = columns_C172[2]
                                cont_C172['ALIQ_ISSQN'] = columns_C172[3]
                                cont_C172['VL_ISSQN'] = columns_C172[4]
                                row_C172 = pd.DataFrame(cont_C172, index=[0])
                                df_C172 = pd.concat([df_C172, row_C172])
                            elif '|C173|' in first_six_chars:
                                columns_C173 = line.strip().split('|')
                                cont_C173['REG'] = columns_C173[1]
                                cont_C173['LOTE_MED'] = columns_C173[2]
                                cont_C173['QTD_ITEM'] = columns_C173[3]
                                cont_C173['DT_FAB'] = columns_C173[4]
                                cont_C173['DT_VAL'] = columns_C173[5]
                                cont_C173['IND_MED'] = columns_C173[6]
                                cont_C173['TP_PROD'] = columns_C173[7]
                                cont_C173['VL_TAB_MAX'] = columns_C173[8]
                                row_C173 = pd.DataFrame(cont_C173, index=[0])
                                df_C173 = pd.concat([df_C173, row_C173])
                            elif '|C174|' in first_six_chars:
                                columns_C174 = line.strip().split('|')
                                cont_C174['REG'] = columns_C174[1]
                                cont_C174['IND_ARM'] = columns_C174[2]
                                cont_C174['NUM_ARM'] = columns_C174[3]
                                cont_C174['DESCR_COMPL'] = columns_C174[4]
                                row_C174 = pd.DataFrame(cont_C174, index=[0])
                                df_C174 = pd.concat([df_C174, row_C174])
                            elif '|C175|' in first_six_chars:
                                columns_C175 = line.strip().split('|')
                                cont_C175['REG'] = columns_C175[1]
                                cont_C175['IND_VEIC_OPER'] = columns_C175[2]
                                cont_C175['CNPJ'] = columns_C175[3]
                                cont_C175['UF'] = columns_C175[4]
                                cont_C175['CHASSI_VEIC'] = columns_C175[5]
                                row_C175 = pd.DataFrame(cont_C175, index=[0])
                                df_C175 = pd.concat([df_C175, row_C175])
                            elif '|C176|' in first_six_chars:
                                columns_C176 = line.strip().split('|')
                                cont_C176['REG'] = columns_C176[1]
                                cont_C176['COD_MOD_ULT_E'] = columns_C176[2]
                                cont_C176['NUM_DOC_ULT_E'] = columns_C176[3]
                                cont_C176['SER_ULT_E'] = columns_C176[4]
                                cont_C176['DT_ULT_E'] = columns_C176[5]
                                cont_C176['COD_PART_ULT_E'] = columns_C176[6]
                                cont_C176['QUANT_ULT_E'] = columns_C176[7]
                                cont_C176['VL_UNIT_ULT_E'] = columns_C176[8]
                                cont_C176['VL_UNIT_BC_ST'] = columns_C176[9]
                                cont_C176['CHAVE_NFE_ULT_E'] = columns_C176[10]
                                cont_C176['NUM_ITEM_ULT_E'] = columns_C176[11]
                                cont_C176['VL_UNIT_BC_ICMS_ULT_E'] = columns_C176[12]
                                cont_C176['ALIQ_ICMS_ULT_E'] = columns_C176[13]
                                cont_C176['VL_UNIT_LIMITE_B_C_ICMS_ULT_E'] = columns_C176[14]
                                cont_C176['VL_UNIT_ICMS_UL_T_E'] = columns_C176[15]
                                cont_C176['ALIQ_ST_ULT_E'] = columns_C176[16]
                                cont_C176['VL_UNIT_RES'] = columns_C176[17]
                                cont_C176['COD_RESP_RET'] = columns_C176[18]
                                cont_C176['COD_MOT_RES'] = columns_C176[19]
                                cont_C176['CHAVE_NFE_RET'] = columns_C176[20]
                                cont_C176['COD_PART_NFE_RET'] = columns_C176[21]
                                cont_C176['SER_NFE_RET'] = columns_C176[22]
                                cont_C176['NUM_NFE_RET'] = columns_C176[23]
                                cont_C176['ITEM_NFE_RET'] = columns_C176[24]
                                cont_C176['COD_DA'] = columns_C176[25]
                                cont_C176['NUM_DA'] = columns_C176[26]
                                cont_C176['VL_UNIT_RES_FCP_ST'] = columns_C176[27]
                                row_C176 = pd.DataFrame(cont_C176, index=[0])
                                df_C176 = pd.concat([df_C176, row_C176])
                            elif '|C177|' in first_six_chars:
                                columns_C177 = line.strip().split('|')
                                cont_C177['REG'] = columns_C177[1]
                                cont_C177['COD_SELO_IPI'] = columns_C177[2]#####################
                                cont_C177['QT_SELO_IPI'] = columns_C177[3]
                                row_C177 = pd.DataFrame(cont_C177, index=[0])
                                df_C177 = pd.concat([df_C177, row_C177])
                            elif '|C178|' in first_six_chars:
                                columns_C178 = line.strip().split('|')
                                cont_C178['REG'] = columns_C178[1]
                                cont_C178['CL_ENQ'] = columns_C178[2]
                                cont_C178['VL_UNID'] = columns_C178[3]
                                cont_C178['QUANT_PAD'] = columns_C178[4]
                                row_C178 = pd.DataFrame(cont_C178, index=[0])
                                df_C178 = pd.concat([df_C178, row_C178])
                            elif '|C179|' in first_six_chars:
                                columns_C179 = line.strip().split('|')
                                cont_C179['REG'] = columns_C179[1]
                                cont_C179['BC_ST_ORIG_DEST'] = columns_C179[2]
                                cont_C179['ICMS_ST_REP'] = columns_C179[3]
                                cont_C179['ICMS_ST_COMPL'] = columns_C179[4]
                                cont_C179['BC_RET'] = columns_C179[5]
                                cont_C179['ICMS_RET'] = columns_C179[6]
                                row_C179 = pd.DataFrame(cont_C179, index=[0])
                                df_C179 = pd.concat([df_C179, row_C179])
                            elif '|C180|' in first_six_chars:
                                columns_C180 = line.strip().split('|')
                                cont_C180['REG'] = columns_C180[1]
                                cont_C180['COD_RESP_RET'] = columns_C180[2]
                                cont_C180['QUANT_CONV'] = columns_C180[3]
                                cont_C180['UNID'] = columns_C180[4]
                                cont_C180['VL_UNIT_CONV'] = columns_C180[5]
                                cont_C180['VL_UNIT_ICMS_OP_CONV'] = columns_C180[6]
                                cont_C180['VL_UNIT_BC_ICMS_ST_CONV'] = columns_C180[7]
                                cont_C180['VL_UNIT_ICMS_ST_CONV'] = columns_C180[8]
                                cont_C180['VL_UNIT_FCP_ST_CONV'] = columns_C180[9]
                                cont_C180['COD_DA'] = columns_C180[10]
                                cont_C180['NUM_DA'] = columns_C180[11]
                                row_C180 = pd.DataFrame(cont_C180, index=[0])
                                df_C180 = pd.concat([df_C180, row_C180])
                            elif '|C181|' in first_six_chars:
                                columns_C181 = line.strip().split('|')
                                cont_C181['REG'] = columns_C181[1]
                                cont_C181['COD_MOT_REST_COMPL'] = columns_C181[2]
                                cont_C181['QUANT_CONV'] = columns_C181[3]
                                cont_C181['UNID'] = columns_C181[4]
                                cont_C181['COD_MOD_SAIDA'] = columns_C181[5]
                                cont_C181['SERIE_SAIDA'] = columns_C181[6]
                                cont_C181['ECF_FAB_SAIDA'] = columns_C181[7]
                                cont_C181['NUM_DOC_SAIDA'] = columns_C181[8]
                                cont_C181['CHV_DFE_SAIDA'] = columns_C181[9]
                                cont_C181['DT_DOC_SAIDA'] = columns_C181[10]
                                cont_C181['NUM_ITEM_SAIDA'] = columns_C181[11]
                                cont_C181['VL_UNIT_CONV_SAIDA'] = columns_C181[12]
                                cont_C181['VL_UNIT_ICMS_OP_ESTOQUE_CONV_SAIDA'] = columns_C181[13]
                                cont_C181['VL_UNIT_ICMS_ST_ESTOQUE_CONV_SAIDA'] = columns_C181[14]
                                cont_C181['VL_UNIT_FCP_ICMS_ST_ESTOQUE_CONV_SAIDA'] = columns_C181[15]
                                cont_C181['VL_UNIT_ICMS_NA_OPERACAO_CONV_SAIDA'] = columns_C181[16]
                                cont_C181['VL_UNIT_ICMS_OP_CONV_SAIDA'] = columns_C181[17]
                                cont_C181['VL_UNIT_ICMS_ST_CONV_REST'] = columns_C181[18]
                                cont_C181['VL_UNIT_FCP_ST_CONV_REST'] = columns_C181[19]
                                cont_C181['VL_UNIT_ICMS_ST_CONV_COMPL'] = columns_C181[20]
                                cont_C181['VL_UNIT_FCP_ST_CONV_COMPL'] = columns_C181[21]
                                row_C181 = pd.DataFrame(cont_C181, index=[0])
                                df_C181 = pd.concat([df_C181, row_C181])
                            elif '|C185|' in first_six_chars:
                                columns_C185 = line.strip().split('|')
                                cont_C185['REG'] = columns_C185[1]
                                cont_C185['NUM_ITEM'] = columns_C185[2]
                                cont_C185['COD_ITEM'] = columns_C185[3]
                                cont_C185['CST_ICMS'] = columns_C185[4]
                                cont_C185['CFOP'] = columns_C185[5]
                                cont_C185['COD_MOT_REST_COMPL'] = columns_C185[6]
                                cont_C185['QUANT_CONV'] = columns_C185[7]
                                cont_C185['UNID'] = columns_C185[8]
                                cont_C185['VL_UNIT_CONV'] = columns_C185[9]
                                cont_C185['VL_UNIT_ICMS_NA_OPERACAO_CONV'] = columns_C185[10]
                                cont_C185['VL_UNIT_ICMS_OP_CONV'] = columns_C185[11]
                                cont_C185['VL_UNIT_ICMS_OP_ESTOQUE_CONV'] = columns_C185[12]
                                cont_C185['VL_UNIT_ICMS_ST_ESTOQUE_CONV'] = columns_C185[13]
                                cont_C185['VL_UNIT_FCP_ICMS_ST_ESTOQUE_CONV'] = columns_C185[14]
                                cont_C185['VL_UNIT_ICMS_ST_CONV_REST'] = columns_C185[15]
                                cont_C185['VL_UNIT_FCP_ST_CONV_REST'] = columns_C185[16]
                                cont_C185['VL_UNIT_ICMS_ST_CONV_COMPL'] = columns_C185[17]
                                cont_C185['VL_UNIT_FCP_ST_CONV_COMPL'] = columns_C185[18]
                                row_C185 = pd.DataFrame(cont_C185, index=[0])
                                df_C185 = pd.concat([df_C185, row_C185])



                except FileNotFoundError as e:
                    print(f"Arquivo não encontrado: {path}")
                    return
                except PermissionError as e:
                    print(f"Erro de permissão ao acessar o arquivo: {path}")
                    return
                except IOError as e:
                    print(f"Erro de E/S ao abrir o arquivo: {path}")
                    return

                pbar.update(1)  # Atualiza a barra de progresso
                update_progress(pbar.n)  # Atualiza a barra de progresso da janela

    # Mostrar mensagem de conclusão
    show_completion_message()

    # Fechar a janela
    root.destroy()

    if dado == '1':
        # if dado equal to 1 I need to send the df_0000 and df_0150 to parte_pandas
        from ..main import MyApp
        app = MyApp.get_running_app()
        get_all_df = [df_0000, df_0001, df_0002, df_0005, df_0015, df_0100, df_0150, df_0175, df_0190, df_0200, df_0205]
        # 0 = efd icms, 1 = efd pis, 2 = ecd icms, 3 = ecd pis
        type = '0'
        app.receive_df(get_all_df, type)
        return True
    else:
        for r in dataframe_to_rows(df_0000, index=False, header=True):
            ws_a.append(r)

        for r in dataframe_to_rows(df_0150, index=False, header=True):
            ws_b.append(r)

        first_sheet = wb.sheetnames[0]
        if first_sheet != ws_a.title and first_sheet != ws_b.title:
            sheet = wb[first_sheet]
            wb.remove(sheet)

        wb.save("D:\\Documents\\app\\test_efd_ICMS.xlsx")
