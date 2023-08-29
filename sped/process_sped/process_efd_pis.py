import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm
from tkinter import Tk, ttk
from tkinter.messagebox import showinfo

def EFD_PIS(pis_path,dado,batch_size=20):
    data = ''

    def FormatarData(variavel, num):
        global data
        data = variavel[num][:2] + '/' + variavel[num][2:4] + '/' + variavel[num][4:]
        return data

    pis_list = pis_path

    # Criar um workbook vazio
    wb = Workbook()

    # Criar duas planilhas vazias
    ws_a = wb.create_sheet("0000")
    ws_b = wb.create_sheet("0150")

    # Criar dataframes vazios
    # Bloco 0
    df_0000 = pd.DataFrame()
    df_0001 = pd.DataFrame()
    df_0035 = pd.DataFrame()
    df_0100 = pd.DataFrame()
    df_0110 = pd.DataFrame()
    df_0111 = pd.DataFrame()
    df_0120 = pd.DataFrame()
    df_0140 = pd.DataFrame()
    df_0145 = pd.DataFrame()
    df_0150 = pd.DataFrame()
    df_0190 = pd.DataFrame()
    df_0200 = pd.DataFrame()
    df_0205 = pd.DataFrame()
    df_0206 = pd.DataFrame()
    df_0208 = pd.DataFrame()
    df_0400 = pd.DataFrame()
    df_0450 = pd.DataFrame()
    df_0500 = pd.DataFrame()
    df_0600 = pd.DataFrame()
    df_0900 = pd.DataFrame()
    df_0990 = pd.DataFrame()
    #Bloco A
    df_A001 = pd.DataFrame()
    df_A010 = pd.DataFrame()
    df_A100 = pd.DataFrame()
    df_A110 = pd.DataFrame()
    df_A111 = pd.DataFrame()
    df_A120 = pd.DataFrame()
    df_A170 = pd.DataFrame()
    df_A990 = pd.DataFrame()
    #Bloco C
    df_C001 = pd.DataFrame()
    df_C010 = pd.DataFrame()
    df_C100 = pd.DataFrame()
    df_C110 = pd.DataFrame()
    df_C111 = pd.DataFrame()
    df_C120 = pd.DataFrame()
    df_C170 = pd.DataFrame()
    df_C175 = pd.DataFrame()
    df_C180 = pd.DataFrame()
    df_C181 = pd.DataFrame()
    df_C185 = pd.DataFrame()
    df_C188 = pd.DataFrame()
    df_C190 = pd.DataFrame()
    df_C191 = pd.DataFrame()
    df_C195 = pd.DataFrame()
    df_C198 = pd.DataFrame()
    df_C199 = pd.DataFrame()
    df_C380 = pd.DataFrame()
    df_C381 = pd.DataFrame()
    df_C385 = pd.DataFrame()
    df_C395 = pd.DataFrame()
    df_C396 = pd.DataFrame()
    df_C400 = pd.DataFrame()
    df_C405 = pd.DataFrame()
    df_C481 = pd.DataFrame()
    df_C485 = pd.DataFrame()
    df_C489 = pd.DataFrame()
    df_C490 = pd.DataFrame()
    df_C491 = pd.DataFrame()
    df_C495 = pd.DataFrame()
    df_C499 = pd.DataFrame()
    df_C500 = pd.DataFrame()
    df_C501 = pd.DataFrame()
    df_C505 = pd.DataFrame()
    df_C509 = pd.DataFrame()
    df_C600 = pd.DataFrame()
    df_C601 = pd.DataFrame()
    df_C605 = pd.DataFrame()
    df_C609 = pd.DataFrame()
    df_C800 = pd.DataFrame()
    df_C810 = pd.DataFrame()
    df_C810 = pd.DataFrame()
    df_C820 = pd.DataFrame()
    df_C830 = pd.DataFrame()
    df_C860 = pd.DataFrame()
    df_C870 = pd.DataFrame()
    df_C880 = pd.DataFrame()
    df_C890 = pd.DataFrame()
    df_C990 = pd.DataFrame()
    #Bloco D
    df_D001 = pd.DataFrame()
    df_D010 = pd.DataFrame()
    df_D100 = pd.DataFrame()
    df_D101 = pd.DataFrame()
    df_D105 = pd.DataFrame()
    df_D111 = pd.DataFrame()
    df_D200 = pd.DataFrame()
    df_D201 = pd.DataFrame()
    df_D205 = pd.DataFrame()
    df_D209 = pd.DataFrame()
    df_D300 = pd.DataFrame()
    df_D309 = pd.DataFrame()
    df_D350 = pd.DataFrame()
    df_D359 = pd.DataFrame()
    df_D500 = pd.DataFrame()
    df_D501 = pd.DataFrame()
    df_D505 = pd.DataFrame()
    df_D509 = pd.DataFrame()
    df_D600 = pd.DataFrame()
    df_D601 = pd.DataFrame()
    df_D605 = pd.DataFrame()
    df_D609 = pd.DataFrame()
    df_D990 = pd.DataFrame()
    #Bloco F
    df_F001 = pd.DataFrame()
    df_F010 = pd.DataFrame()
    df_F100 = pd.DataFrame()
    df_F111 = pd.DataFrame()
    df_F120 = pd.DataFrame()
    df_F129 = pd.DataFrame()
    df_F130 = pd.DataFrame()
    df_F139 = pd.DataFrame()
    df_F150 = pd.DataFrame()
    df_F200 = pd.DataFrame()
    df_F205 = pd.DataFrame()
    df_F210 = pd.DataFrame()
    df_F211 = pd.DataFrame()
    df_F500 = pd.DataFrame()
    df_F509 = pd.DataFrame()
    df_F510 = pd.DataFrame()
    df_F519 = pd.DataFrame()
    df_F525 = pd.DataFrame()
    df_F550 = pd.DataFrame()
    df_F559 = pd.DataFrame()
    df_F560 = pd.DataFrame()
    df_F569 = pd.DataFrame()
    df_F600 = pd.DataFrame()
    df_F700 = pd.DataFrame()
    df_F800 = pd.DataFrame()
    df_F990 = pd.DataFrame()
    #Bloco I
    df_I001 = pd.DataFrame()
    df_I010 = pd.DataFrame()
    df_I100 = pd.DataFrame()
    df_I199 = pd.DataFrame()
    df_I200 = pd.DataFrame()
    df_I299 = pd.DataFrame()
    df_I300 = pd.DataFrame()
    df_I399 = pd.DataFrame()
    df_I990 = pd.DataFrame()
    #Bloco M
    df_M001 = pd.DataFrame()
    df_M100 = pd.DataFrame()
    df_M105 = pd.DataFrame()
    df_M110 = pd.DataFrame()
    df_M115 = pd.DataFrame()
    df_M200 = pd.DataFrame()
    df_M205 = pd.DataFrame()
    df_M210 = pd.DataFrame()
    df_M211 = pd.DataFrame()
    df_M215 = pd.DataFrame()
    df_M220 = pd.DataFrame()
    df_M225 = pd.DataFrame()
    df_M230 = pd.DataFrame()
    df_M300 = pd.DataFrame()
    df_M350 = pd.DataFrame()
    df_M400 = pd.DataFrame()
    df_M410 = pd.DataFrame()
    df_M500 = pd.DataFrame()
    df_M505 = pd.DataFrame()
    df_M510 = pd.DataFrame()
    df_M515 = pd.DataFrame()
    df_M600 = pd.DataFrame()
    df_M605 = pd.DataFrame()
    df_M610 = pd.DataFrame()
    df_M611 = pd.DataFrame()
    df_M615 = pd.DataFrame()
    df_M620 = pd.DataFrame()
    df_M625 = pd.DataFrame()
    df_M630 = pd.DataFrame()
    df_M700 = pd.DataFrame()
    df_M800 = pd.DataFrame()
    df_M810 = pd.DataFrame()
    df_M990 = pd.DataFrame()
    #Bloco P
    df_P001 = pd.DataFrame()
    df_P010 = pd.DataFrame()
    df_P100 = pd.DataFrame()
    df_P110 = pd.DataFrame()
    df_P199 = pd.DataFrame()
    df_P200 = pd.DataFrame()
    df_P210 = pd.DataFrame()
    df_P990 = pd.DataFrame()
    #Bloco 1
    df_1001 = pd.DataFrame()
    df_1010 = pd.DataFrame()
    df_1011 = pd.DataFrame()
    df_1020 = pd.DataFrame()
    df_1050 = pd.DataFrame()
    df_1100 = pd.DataFrame()
    df_1101 = pd.DataFrame()
    df_1102 = pd.DataFrame()
    df_1200 = pd.DataFrame()
    df_1210 = pd.DataFrame()
    df_1220 = pd.DataFrame()
    df_1300 = pd.DataFrame()
    df_1500 = pd.DataFrame()
    df_1501 = pd.DataFrame()
    df_1502 = pd.DataFrame()
    df_1600 = pd.DataFrame()
    df_1610 = pd.DataFrame()
    df_1620 = pd.DataFrame()
    df_1700 = pd.DataFrame()
    df_1800 = pd.DataFrame()
    df_1809 = pd.DataFrame()
    df_1900 = pd.DataFrame()
    df_1990 = pd.DataFrame()
    #Bloco 9
    df_9001 = pd.DataFrame()
    df_9900 = pd.DataFrame()
    df_9990 = pd.DataFrame()
    df_9999 = pd.DataFrame()

    df_0150 = pd.DataFrame()
    df_xxx = pd.DataFrame()

    # Função para atualizar a barra de progresso
    def update_progress(count):
        progress_bar["value"] = count
        root.update_idletasks()

    # Função para mostrar mensagem ao concluir
    def show_completion_message():
        showinfo("Concluído", "Processamento concluído com sucesso!")

    # Configurar janela
    root = Tk()
    root.title("Processando arquivos")
    root.geometry("300x100")

    # Criar barra de progresso
    progress_bar = ttk.Progressbar(root, length=200, mode="determinate")
    progress_bar.pack(pady=20)

    # Percorrer a lista de diretórios/arquivos em lotes com barra de progresso
    with tqdm(total=len(pis_list), desc="Processando arquivos", bar_format="{desc}") as pbar:
        for i in range(0, len(pis_list), batch_size):
            batch_paths = pis_list[i:i+batch_size]

            for path in batch_paths:
                try:
                    replace_pis_str = path.replace("/", "\\")

                    with open(replace_pis_str, 'r', encoding="latin-1") as file:
                        # Bloco 0
                        cont_0000 = {}
                        cont_0001 = {}
                        cont_0035 = {}
                        cont_0100 = {}
                        cont_0110 = {}
                        cont_0111 = {}
                        cont_0120 = {}
                        cont_0140 = {}
                        cont_0145 = {}
                        cont_0150 = {}
                        cont_0190 = {}
                        cont_0200 = {}
                        cont_0205 = {}
                        cont_0206 = {}
                        cont_0208 = {}
                        cont_0400 = {}
                        cont_0450 = {}
                        cont_0500 = {}
                        cont_0600 = {}
                        cont_0900 = {}
                        cont_0990 = {}
                        # Bloco A
                        cont_A001 = {}
                        cont_A010 = {}
                        cont_A100 = {}
                        cont_A110 = {}
                        cont_A111 = {}
                        cont_A120 = {}
                        cont_A170 = {}
                        cont_A990 = {}
                        # Bloco C
                        cont_C001 = {}
                        cont_C010 = {}
                        cont_C100 = {}
                        cont_C110 = {}
                        cont_C111 = {}
                        cont_C120 = {}
                        cont_C170 = {}
                        cont_C175 = {}
                        cont_C180 = {}
                        cont_C181 = {}
                        cont_C185 = {}
                        cont_C188 = {}
                        cont_C190 = {}
                        cont_C191 = {}
                        cont_C195 = {}
                        cont_C198 = {}
                        cont_C199 = {}
                        cont_C380 = {}
                        cont_C381 = {}
                        cont_C385 = {}
                        cont_C395 = {}
                        cont_C396 = {}
                        cont_C400 = {}
                        cont_C405 = {}
                        cont_C481 = {}
                        cont_C485 = {}
                        cont_C489 = {}
                        cont_C490 = {}
                        cont_C491 = {}
                        cont_C495 = {}
                        cont_C499 = {}
                        cont_C500 = {}
                        cont_C501 = {}
                        cont_C505 = {}
                        cont_C509 = {}
                        cont_C600 = {}
                        cont_C601 = {}
                        cont_C605 = {}
                        cont_C609 = {}
                        cont_C800 = {}
                        cont_C810 = {}
                        cont_C810 = {}
                        cont_C820 = {}
                        cont_C830 = {}
                        cont_C860 = {}
                        cont_C870 = {}
                        cont_C880 = {}
                        cont_C890 = {}
                        cont_C990 = {}
                        # Bloco D
                        cont_D001 = {}
                        cont_D010 = {}
                        cont_D100 = {}
                        cont_D101 = {}
                        cont_D105 = {}
                        cont_D111 = {}
                        cont_D200 = {}
                        cont_D201 = {}
                        cont_D205 = {}
                        cont_D209 = {}
                        cont_D300 = {}
                        cont_D309 = {}
                        cont_D350 = {}
                        cont_D359 = {}
                        cont_D500 = {}
                        cont_D501 = {}
                        cont_D505 = {}
                        cont_D509 = {}
                        cont_D600 = {}
                        cont_D601 = {}
                        cont_D605 = {}
                        cont_D609 = {}
                        cont_D990 = {}
                        # Bloco F
                        cont_F001 = {}
                        cont_F010 = {}
                        cont_F100 = {}
                        cont_F111 = {}
                        cont_F120 = {}
                        cont_F129 = {}
                        cont_F130 = {}
                        cont_F139 = {}
                        cont_F150 = {}
                        cont_F200 = {}
                        cont_F205 = {}
                        cont_F210 = {}
                        cont_F211 = {}
                        cont_F500 = {}
                        cont_F509 = {}
                        cont_F510 = {}
                        cont_F519 = {}
                        cont_F525 = {}
                        cont_F550 = {}
                        cont_F559 = {}
                        cont_F560 = {}
                        cont_F569 = {}
                        cont_F600 = {}
                        cont_F700 = {}
                        cont_F800 = {}
                        cont_F990 = {}
                        # Bloco I
                        cont_I001 = {}
                        cont_I010 = {}
                        cont_I100 = {}
                        cont_I199 = {}
                        cont_I200 = {}
                        cont_I299 = {}
                        cont_I300 = {}
                        cont_I399 = {}
                        cont_I990 = {}
                        # Bloco M
                        cont_M001 = {}
                        cont_M100 = {}
                        cont_M105 = {}
                        cont_M110 = {}
                        cont_M115 = {}
                        cont_M200 = {}
                        cont_M205 = {}
                        cont_M210 = {}
                        cont_M211 = {}
                        cont_M215 = {}
                        cont_M220 = {}
                        cont_M225 = {}
                        cont_M230 = {}
                        cont_M300 = {}
                        cont_M350 = {}
                        cont_M400 = {}
                        cont_M410 = {}
                        cont_M500 = {}
                        cont_M505 = {}
                        cont_M510 = {}
                        cont_M515 = {}
                        cont_M600 = {}
                        cont_M605 = {}
                        cont_M610 = {}
                        cont_M611 = {}
                        cont_M615 = {}
                        cont_M620 = {}
                        cont_M625 = {}
                        cont_M630 = {}
                        cont_M700 = {}
                        cont_M800 = {}
                        cont_M810 = {}
                        cont_M990 = {}
                        # Bloco P
                        cont_P001 = {}
                        cont_P010 = {}
                        cont_P100 = {}
                        cont_P110 = {}
                        cont_P199 = {}
                        cont_P200 = {}
                        cont_P210 = {}
                        cont_P990 = {}
                        # Bloco 1
                        cont_1001 = {}
                        cont_1010 = {}
                        cont_1011 = {}
                        cont_1020 = {}
                        cont_1050 = {}
                        cont_1100 = {}
                        cont_1101 = {}
                        cont_1102 = {}
                        cont_1200 = {}
                        cont_1210 = {}
                        cont_1220 = {}
                        cont_1300 = {}
                        cont_1500 = {}
                        cont_1501 = {}
                        cont_1502 = {}
                        cont_1600 = {}
                        cont_1610 = {}
                        cont_1620 = {}
                        cont_1700 = {}
                        cont_1800 = {}
                        cont_1809 = {}
                        cont_1900 = {}
                        cont_1990 = {}
                        # Bloco 9
                        cont_9001 = {}
                        cont_9900 = {}
                        cont_9990 = {}
                        cont_9999 = {}

                        for line in file:
                            first_six_chars = line[:6]

                            if '|0000|' in first_six_chars:
                                # Split the line by '|' and store the resulting columns in a list
                                columns_0000 = line.strip().split('|')
                                # assign to every column
                                #cont_0000['ID_DT_INI'] = data
                                #FormatarData(columns_0000, 7)
                                # cont_0000['ID_CNPJ'] = columns_0000.get(9, '') para teste caso for nulo futuramente
                                cont_0000['REG'] = columns_0000[1]
                                cont_0000['COD_VER'] = columns_0000[2]
                                cont_0000['TIPO_ESCRIT'] = columns_0000[3]
                                cont_0000['IND_SIT_ESP'] = columns_0000[4]
                                cont_0000['NUM_REC_ANTERIOR'] = columns_0000[5]
                                cont_0000['DT_INI'] = columns_0000[6]
                                cont_0000['DT_FIN'] = columns_0000[7]
                                cont_0000['NOME'] = columns_0000[8]
                                cont_0000['CNPJ'] = columns_0000[9]
                                cont_0000['UF'] = columns_0000[10]
                                cont_0000['COD_MUN'] = columns_0000[11]
                                cont_0000['SUFRAMA'] = columns_0000[12]
                                cont_0000['IND_NAT_PJ'] = columns_0000[13]
                                cont_0000['IND_ATIV'] = columns_0000[14]
                                # Create a new row in the dataframe using the columns list as the values
                                row_0000 = pd.DataFrame(cont_0000, index=[0])
                                # Concatenate the row to the dataframe
                                df_0000 = pd.concat([df_0000, row_0000])
                            elif '|0035|' in first_six_chars:
                                # Split the line by '|' and store the resulting columns in a list
                                columns_0035 = line.strip().split('|')
                                cont_0035['REG'] = columns_0035[1]
                                cont_0035['COD_SCP'] = columns_0035[2]
                                cont_0035['DESC_SCP'] = columns_0035[3]
                                cont_0035['INF_COMP'] = columns_0035[4]
                                row_0035 = pd.DataFrame(cont_0035, index=[0])
                                df_0035 = pd.concat([df_0035, row_0035])
                            elif '|0100|' in first_six_chars:
                                columns_0100 = line.strip().split('|')
                                cont_0100['REG'] = columns_0100[1]
                                cont_0100['NOME'] = columns_0100[2]
                                cont_0100['CPF'] = columns_0100[3]
                                cont_0100['CRC'] = columns_0100[4]
                                cont_0100['CNPJ'] = columns_0100[5]
                                cont_0100['CEP'] = columns_0100[6]
                                cont_0100['END'] = columns_0100[7]
                                cont_0100['NUM'] = columns_0100[8]
                                cont_0100['COMPL'] = columns_0100[9]
                                cont_0100['BAIRRO'] = columns_0100[10]
                                cont_0100['FONE'] = columns_0100[11]
                                cont_0100['FAX'] = columns_0100[12]
                                cont_0100['EMAIL'] = columns_0100[13]
                                cont_0100['COD_MUN'] = columns_0100[14]
                                row_0100 = pd.DataFrame(cont_0100, index=[0])
                                df_0100 = pd.concat([df_0100, row_0100])

                except FileNotFoundError as e:
                    print(f"Arquivo não encontrado: {path}")
                    return
                except PermissionError as e:
                    print(f"Erro de permissão ao acessar o arquivo: {path}")
                    return
                except IOError as e:
                    print(f"Erro de E/S ao abrir o arquivo: {path}")
                    return

                pbar.update(1)  # Atualiza a barra de progresso
                update_progress(pbar.n)  # Atualiza a barra de progresso da janela

    # Mostrar mensagem de conclusão
    show_completion_message()

    # Fechar a janela
    root.destroy()

    if dado == '1':
        #if dado equal to 1 I need to send the df_0000 and df_0150 to parte_pandas
        from ..main import MyApp
        app = MyApp.get_running_app()
        get_all_df = [df_0000,df_0035,df_0100,df_C170]
        # 0 = efd icms, 1 = efd pis, 2 = ecd icms, 3 = ecd pis
        type = '1'
        app.receive_df(get_all_df,type)
        return True
    else:#else export it as excel csv
        for r in dataframe_to_rows(df_0000, index=False, header=True):
            ws_a.append(r)

        for r in dataframe_to_rows(df_0150, index=False, header=True):
            ws_b.append(r)

        first_sheet = wb.sheetnames[0]
        if first_sheet != ws_a.title and first_sheet != ws_b.title:
            sheet = wb[first_sheet]
            wb.remove(sheet)

        wb.save("C:\\Documents\\app\\test_efd_PIS.xlsx")
